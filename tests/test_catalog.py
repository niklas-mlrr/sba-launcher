"""Tests für ``core.catalog`` — Katalog-Modell, Excel-Import/Render, Overrides.

tkinter-frei und ohne IServ-Kontakt: arbeitet gegen ein synthetisches Excel-
Fixture im tmp-Verzeichnis (dasselbe Layout wie die echte Bestand-Excel: Fach-/
Zustand-/Jahrgang-Zeilen + merged Mehrjahresband-Zellen + ein ``zu Bestellen``-
Sheet). Keine Abhängigkeit von der realen 2026er Datei.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from core import catalog

# ISBNs im Fixture (plain digits).
_ISBN_DEU5 = "9783062052224"
_ISBN_DEU6 = "9783062052231"
_ISBN_ERDE_MJB = "9783121052073"
_ISBN_PHYSIK = "9783123456789"  # ohne Layout-Slot (für unmatched-Test)


def _build_fixture(tmp_path: Path) -> Path:
    """Baut ein kleines Bestand-Workbook: Deutsch (Jg5/6 single) + Erdkunde (MJB 5-6)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Bestand- und Nachbestellung"

    # Zeile 1: Fach-Header (merged über die 4 Zustand-Spalten je Buch).
    ws["A1"] = "Fach"
    ws["B1"] = "Deutsch"
    ws.merge_cells("B1:E1")
    ws["F1"] = "Erdkunde"
    ws.merge_cells("F1:I1")

    # Zeile 2: Zustand-Header.
    ws["A2"] = "Zustand"
    _zustaende = ["Angemeldet", "Bestand", "Bestellt", "zu bestellen"]
    for col, label in zip("BCDE", _zustaende, strict=False):
        ws[f"{col}2"] = label
    for col, label in zip("FGHI", _zustaende, strict=False):
        ws[f"{col}2"] = label

    # Zeilen 3-4: Jahrgänge 5 und 6 (+ Dummy-Counts).
    ws["A3"] = "Jahrgang 5"
    ws["A4"] = "Jahrgang 6"
    ws["B3"], ws["C3"] = 10, 11
    ws["B4"], ws["C4"] = 12, 13
    ws["F3"], ws["G3"] = 5, 6
    ws["F4"], ws["G4"] = 7, 8

    # Erdkunde als Mehrjahresband: Angemeldet + Bestand über Jg5+Jg6 merged.
    ws.merge_cells("F3:F4")
    ws.merge_cells("G3:G4")

    # Sheet 'zu Bestellen' (nur Deutsch 5 hat Nachbestellbedarf → Anreicherung).
    wz = wb.create_sheet("zu Bestellen")
    wz["B1"], wz["C1"], wz["E1"] = "Jahrgang", "Fach", "Verfasser/Titel"
    wz["F1"], wz["G1"], wz["H1"] = "Verlag", "Best.-Nr. (ISBN)", "Einzelpreis (brutto)"
    wz["B2"], wz["C2"], wz["E2"] = 5, "Deutsch", "Deutschbuch 5"
    wz["F2"], wz["G2"], wz["H2"] = "Cornelsen", "978-3-06-205222-4", 30.99

    path = tmp_path / "fixture.xlsx"
    wb.save(path)
    return path


def _mappings_deu_erde() -> list[dict]:
    return [
        {"isbn": _ISBN_DEU5, "bestand_cell": "C3"},
        {"isbn": _ISBN_DEU6, "bestand_cell": "C4"},
        {"isbn": _ISBN_ERDE_MJB, "bestand_cell": "G3"},  # merged G3:G4
    ]


# --- Modell ---------------------------------------------------------------


def test_mehrjahresband_abgeleitet() -> None:
    single = catalog.Eintrag(fach="Deutsch", jahrgang_von=5, jahrgang_bis=5, isbn=_ISBN_DEU5)
    mjb = catalog.Eintrag(fach="Erdkunde", jahrgang_von=5, jahrgang_bis=6, isbn=_ISBN_ERDE_MJB)
    assert single.mehrjahresband is False
    assert mjb.mehrjahresband is True


def test_id_deterministisch() -> None:
    a = catalog.Eintrag(fach="Deutsch", jahrgang_von=5, jahrgang_bis=5, isbn=_ISBN_DEU5)
    b = catalog.Eintrag(fach="Deutsch", jahrgang_von=5, jahrgang_bis=5, isbn=_ISBN_DEU5)
    assert a.id == b.id and a.id
    # Unterschiedliche Identität → unterschiedliche ID.
    c = catalog.Eintrag(fach="Deutsch", jahrgang_von=6, jahrgang_bis=6, isbn=_ISBN_DEU6)
    assert c.id != a.id


def test_katalog_json_roundtrip(tmp_path: Path) -> None:
    k = catalog.Katalog(
        schule="TRG Osterode",
        schuljahr="2026/2027",
        eintraege=[
            catalog.Eintrag(
                fach="Deutsch", jahrgang_von=5, jahrgang_bis=5, isbn=_ISBN_DEU5, hint=None
            ),
            catalog.Eintrag(fach="Politik", jahrgang_von=11, jahrgang_bis=12, isbn="x", hint="eA"),
        ],
    )
    p = catalog.save_katalog(k, tmp_path / "katalog.json")
    loaded = catalog.load_katalog(p)
    assert loaded.schule == "TRG Osterode"
    assert loaded.schuljahr == "2026/2027"
    assert len(loaded.eintraege) == 2
    assert loaded.eintraege[1].hint == "eA"
    assert loaded.eintraege[1].mehrjahresband is True
    # mehrjahresband wird nicht serialisiert (abgeleitet).
    raw = json.loads(p.read_text(encoding="utf-8"))
    assert "mehrjahresband" not in raw["eintraege"][0]
    # aber beim Laden toleriert (abwärtskompatibel).
    raw["eintraege"][0]["mehrjahresband"] = True
    p.write_text(json.dumps(raw), encoding="utf-8")
    assert catalog.load_katalog(p).eintraege[0].mehrjahresband is False  # von/bis maßgeblich


# --- import_from_excel ----------------------------------------------------


def test_import_from_excel(tmp_path: Path) -> None:
    fx = _build_fixture(tmp_path)
    k = catalog.import_from_excel(
        fx, _mappings_deu_erde(), schule="TRG", schuljahr="2026/2027"
    )
    by_key = {(e.fach, e.jahrgang_von, e.jahrgang_bis): e for e in k.eintraege}
    assert set(by_key) == {("Deutsch", 5, 5), ("Deutsch", 6, 6), ("Erdkunde", 5, 6)}

    deu5 = by_key[("Deutsch", 5, 5)]
    assert deu5.isbn == _ISBN_DEU5
    assert deu5.mehrjahresband is False
    # Anreicherung aus 'zu Bestellen'.
    assert deu5.titel == "Deutschbuch 5"
    assert deu5.verlag == "Cornelsen"
    assert deu5.neupreis == pytest.approx(30.99)

    deu6 = by_key[("Deutsch", 6, 6)]
    assert deu6.isbn == _ISBN_DEU6
    assert deu6.titel == ""  # nicht in 'zu Bestellen'

    erde = by_key[("Erdkunde", 5, 6)]
    assert erde.isbn == _ISBN_ERDE_MJB
    assert erde.mehrjahresband is True
    assert erde.hint is None


# --- catalog_to_overrides -------------------------------------------------


def test_catalog_to_overrides() -> None:
    k = catalog.Katalog(
        schule="", schuljahr="", eintraege=[
            catalog.Eintrag(fach="Deutsch", jahrgang_von=5, jahrgang_bis=5, isbn=_ISBN_DEU5),
            catalog.Eintrag(fach="Erdkunde", jahrgang_von=5, jahrgang_bis=6, isbn=_ISBN_ERDE_MJB),
            catalog.Eintrag(fach="Politik", jahrgang_von=11, jahrgang_bis=12, isbn="px", hint="eA"),
        ]
    )
    ov = catalog.catalog_to_overrides(k)
    assert ov["5|Deutsch|"] == _ISBN_DEU5
    assert ov["5|Erdkunde|"] == _ISBN_ERDE_MJB
    assert ov["6|Erdkunde|"] == _ISBN_ERDE_MJB
    assert ov["11|Politik|eA"] == "px"
    assert ov["12|Politik|eA"] == "px"
    # MJB emittiert pro Jahrgang im Bereich.
    assert sum(1 for key in ov if key.startswith("6|Erdkunde")) == 1


# --- render_excel + Roundtrip ---------------------------------------------


def test_render_excel_mappings(tmp_path: Path) -> None:
    fx = _build_fixture(tmp_path)
    k = catalog.import_from_excel(fx, _mappings_deu_erde())
    out = tmp_path / "out" / "Liste.xlsx"
    out.parent.mkdir()
    config_path, unmatched = catalog.render_excel(fx, k, out)

    assert unmatched == []
    assert config_path == out.parent / "config.json"
    cfg = json.loads(config_path.read_text(encoding="utf-8"))
    assert cfg["excel_file"] == "Liste.xlsx"
    assert cfg["safety_stock"] == 5
    by_isbn = {m["isbn"]: m for m in cfg["mappings"]}
    assert by_isbn[_ISBN_DEU5]["bestand_cell"] == "C3"
    assert by_isbn[_ISBN_DEU6]["bestand_cell"] == "C4"
    assert by_isbn[_ISBN_ERDE_MJB]["bestand_cell"] == "G3"  # merged anchor
    # Overrides aus dem Katalog übernommen.
    assert cfg["match_overrides"]["5|Erdkunde|"] == _ISBN_ERDE_MJB


def test_render_roundtrip(tmp_path: Path) -> None:
    """Import → Katalog → render (Vorlage=Fixture) → re-import → strukturidentisch."""
    fx = _build_fixture(tmp_path)
    k1 = catalog.import_from_excel(fx, _mappings_deu_erde(), schule="TRG", schuljahr="2026/2027")

    out = tmp_path / "out" / "Liste.xlsx"
    out.parent.mkdir()
    config_path, unmatched = catalog.render_excel(fx, k1, out)
    assert unmatched == []

    cfg = json.loads(config_path.read_text(encoding="utf-8"))
    k2 = catalog.import_from_excel(
        out, cfg["mappings"], sheet_name=cfg["sheet_name"], schule="TRG", schuljahr="2026/2027"
    )

    def signature(k: catalog.Katalog) -> list[tuple]:
        return sorted(
            (e.fach, e.hint, e.jahrgang_von, e.jahrgang_bis, e.isbn, e.titel, e.verlag, e.neupreis)
            for e in k.eintraege
        )

    assert signature(k1) == signature(k2)


def test_render_unmatched(tmp_path: Path) -> None:
    fx = _build_fixture(tmp_path)
    k = catalog.import_from_excel(fx, _mappings_deu_erde())
    # Zusätzlicher Eintrag ohne Layout-Slot.
    k.eintraege.append(
        catalog.Eintrag(fach="Physik", jahrgang_von=5, jahrgang_bis=5, isbn=_ISBN_PHYSIK)
    )
    out = tmp_path / "out2" / "Liste.xlsx"
    out.parent.mkdir()
    _config_path, unmatched = catalog.render_excel(fx, k, out)
    assert any("Physik" in u and _ISBN_PHYSIK in u for u in unmatched)
    # Die drei zuordenbaren Einträge landen dennoch in den mappings.
    cfg = json.loads((out.parent / "config.json").read_text(encoding="utf-8"))
    assert len([m for m in cfg["mappings"]]) == 3


# --- load_katalog Fehlertoleranz / atomarer Write --------------------------


def test_load_katalog_corrupt_json_liefert_leer(tmp_path: Path) -> None:
    p = tmp_path / "katalog.json"
    p.write_text("{kein gültiges json", encoding="utf-8")
    k = catalog.load_katalog(p)
    assert k.schule == ""
    assert k.schuljahr == ""
    assert k.eintraege == []


def test_save_katalog_hinterlässt_keine_tmp_datei(tmp_path: Path) -> None:
    k = catalog.Katalog(schule="TRG", schuljahr="2026/2027")
    p = tmp_path / "katalog.json"
    catalog.save_katalog(k, p)
    assert p.is_file()
    # Atomarer Write via tmp+rename → keine Temp-Datei übrig.
    assert not list(tmp_path.glob("*.json.tmp"))
