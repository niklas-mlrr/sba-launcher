"""Katalog-Modell + Excel-Import/Render + match_overrides-Sync (Phase 4).

Der Bestand-Katalog bildet ``Fach × Jahrgang → ISBN`` (plus Mehrjahresband-Flag,
Titel/Verlag/Neupreis). Er ersetzt das manuelle Pflegen von ``config.json:mappings``
und ``match_overrides`` durch einen strukturierten Editor.

Datenheimat: ``data/katalog.json`` (geladen via :func:`load_katalog`).

Excel-Layout (``Bestand- und Nachbestellung``): die Zellen enthalten **Counts**
(Angemeldet/Bestand/Bestellt), keine ISBNs. Die ISBN→Zelle-Zuordnung steht in
``config.json:mappings`` (Liste ``{isbn, bestand_cell, angemeldet_cell?}``). Daher
braucht :func:`import_from_excel` die Excel-Datei **und** die mappings-Liste.

Produktionsschutz (CLAUDE.md): rein lokale Excel-/JSON-IO, kein IServ-Kontakt, keine
API-Writes. isbnlib/openpyxl sind Launcher-Dependencies (s. pyproject.toml), damit
dieses Modul auf dem headless VPS via pytest testbar ist.

tkinter-frei.
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
import shutil
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

from core import paths
from core.config_io import _atomic_write_text

_log = logging.getLogger(__name__)

# LogFn-Callback-Muster (wie core/barcode.py, core/bestand.py, core/ausleihe_ausgabe.py):
# jedes core-Modul hält seinen eigenen Alias statt einen gemeinsamen zu importieren.
LogFn = Callable[[str], None]


def _warn(log: LogFn | None, msg: str) -> None:
    """Warnung an ``log`` (GUI-Konsole) wenn übergeben, sonst an das Modul-Logging.

    ``logging`` wird von der GUI nie konfiguriert, daher landen Warnungen ohne
    injizierten ``log``-Callback unsichtbar auf stderr. Der Fallback bleibt aber
    bestehen, damit Aufrufe ohne ``log`` (z. B. bestehende Tests) nicht brechen.
    """
    if log is not None:
        log(msg)
    else:
        _log.warning(msg)

# ── Layout-Primitiven (gespiegelt aus update_bestand_auto.py) ────────────────
# O5 (2026-08-12): Die reinen Helper werden kopiert statt ``update_bestand_auto``
# zu importieren, weil jenes Modul Top-Level ``from ausleihe import AusleiheClient``
# + ``dotenv.load_dotenv()`` ausführt — ein Import aus dem Launcher-Venv (ohne das
# Paket ``ausleihe``) scheitert bzw. lädt Credentials. Die Kopien halten sich
# byte-identisch zum Original (Quelle: ``sba-bestand/bestand/
# update_bestand_auto.py``); Drift-Risiko akzeptiert (stabil).


def resolve_anchor(ws, row: int, col: int) -> tuple[int, int]:
    """Gibt (anchor_row, anchor_col) zurück – bei Zellenverbund die oben-links-Zelle."""
    cell_ref = f"{get_column_letter(col)}{row}"
    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            return merged.min_row, merged.min_col
    return row, col


def _merged_span(ws, row: int, col: int) -> tuple[int, int]:
    """Zeilen-Spannweite der Zelle: (min_row, max_row) — bei Verbund der volle Bereich."""
    cell_ref = f"{get_column_letter(col)}{row}"
    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            return merged.min_row, merged.max_row
    return row, row


def find_fach_for_col(ws, fach_rows: list[int], col: int) -> str | None:
    """Fach-Label für Spalte col aus der nächsten (untersten) Fach-Zeile mit Inhalt.
    Ist die Zelle leer, wird die nächsthöhere Fach-Zeile als Fallback genommen.
    """
    for fach_row in reversed(fach_rows):
        ar, ac = resolve_anchor(ws, fach_row, col)
        if ar != fach_row:
            # Anker liegt in einer anderen Zeile – diese Zeile überspringen
            continue
        val = ws.cell(ar, ac).value
        if val is not None:
            return str(val)
    return None


def find_zustand_for_col(ws, zustand_rows: list[int], col: int) -> str | None:
    """Zustand-Label für Spalte col, mit Fallback auf höhere Zustand-Zeilen."""
    for zustand_row in reversed(zustand_rows):
        ar, ac = resolve_anchor(ws, zustand_row, col)
        if ar != zustand_row:
            continue
        val = ws.cell(ar, ac).value
        if val is not None:
            return str(val)
    return None


def classify_row(ws, row: int) -> str:
    """'fach' | 'zustand' | 'stand' | 'jahrgang' | 'other'"""
    val = ws.cell(row, 1).value
    if val == "Fach":
        return "fach"
    if val == "Zustand":
        return "zustand"
    if val == "Stand":
        return "stand"
    # Vorlage führt jede Jahrgang-Zeile einzeln (keine "Jahrgang 11/12"-Combined),
    # daher passt die einfache ``Jahrgang \d+``-Regex — Rows müssen ein Jahrgang sein.
    if isinstance(val, str) and re.match(r"Jahrgang\s+\d+", val):
        return "jahrgang"
    return "other"


def extract_grade(ws, row: int) -> int | None:
    val = ws.cell(row, 1).value
    m = re.match(r"Jahrgang\s+(\d+)", str(val)) if val else None
    return int(m.group(1)) if m else None


def strip_hint(text: str) -> tuple[str, str | None]:
    """Trennt Serientitel-Hinweis in Klammern vom Fach-Namen.
    'Politik (eA)' → ('Politik', 'eA'); 'Deutsch' → ('Deutsch', None).
    """
    m = re.match(r"^(.*?)\s*\((.+)\)\s*$", text.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return text.strip(), None


# ── Datenmodell ──────────────────────────────────────────────────────────────


def _make_id(fach: str, hint: str | None, von: int, bis: int, isbn: str) -> str:
    """Deterministische ID — sha1 der Identität (stabil über Seed-Läufe)."""
    raw = f"{fach}|{hint or ''}|{von}|{bis}|{isbn}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


@dataclass
class Eintrag:
    """Ein Katalog-Eintrag: ein Buch für einen Jahrgangs-Bereich in einem Fach."""

    fach: str
    jahrgang_von: int
    jahrgang_bis: int
    isbn: str
    hint: str | None = None
    titel: str = ""
    verlag: str = ""
    neupreis: float = 0.0
    id: str = ""

    def __post_init__(self) -> None:
        if self.jahrgang_von > self.jahrgang_bis:
            # Vertauschtes von/bis (z. B. Tippfehler in der GUI) soll den Eintrag
            # nicht sofort zum Absturz bringen — statt eine Exception zu werfen,
            # wird die Spanne repariert (getauscht) und eine Warnung geloggt.
            # Kein ``log``-Callback verfügbar (Dataclass-Konstruktion) → Fallback
            # auf das Modul-Logging via ``_warn``.
            _warn(
                None,
                f"Eintrag {self.fach!r} [{self.isbn}]: jahrgang_von "
                f"({self.jahrgang_von}) > jahrgang_bis ({self.jahrgang_bis}) "
                "— Werte vertauscht.",
            )
            self.jahrgang_von, self.jahrgang_bis = self.jahrgang_bis, self.jahrgang_von
        if not self.id:
            self.id = _make_id(
                self.fach, self.hint, self.jahrgang_von, self.jahrgang_bis, self.isbn
            )

    @property
    def mehrjahresband(self) -> bool:
        """Abgeleitet: jahrgang_bis > jahrgang_von ⇒ Mehrjahresband."""
        return self.jahrgang_bis > self.jahrgang_von

    def to_json(self) -> dict:
        return {
            "id": self.id,
            "fach": self.fach,
            "hint": self.hint,
            "jahrgang_von": self.jahrgang_von,
            "jahrgang_bis": self.jahrgang_bis,
            "isbn": self.isbn,
            "titel": self.titel,
            "verlag": self.verlag,
            "neupreis": self.neupreis,
        }

    @classmethod
    def from_json(cls, d: dict) -> Eintrag:
        return cls(
            fach=str(d["fach"]),
            jahrgang_von=int(d["jahrgang_von"]),
            jahrgang_bis=int(d["jahrgang_bis"]),
            isbn=str(d["isbn"]),
            hint=d.get("hint"),
            titel=str(d.get("titel", "") or ""),
            verlag=str(d.get("verlag", "") or ""),
            neupreis=float(d.get("neupreis", 0) or 0),
            id=str(d.get("id", "") or ""),
        )


@dataclass
class Katalog:
    """Vollständiger Katalog: Metadaten + Einträge."""

    schule: str
    schuljahr: str
    eintraege: list[Eintrag] = field(default_factory=list)

    def to_json(self) -> dict:
        return {
            "schule": self.schule,
            "schuljahr": self.schuljahr,
            "eintraege": [e.to_json() for e in self.eintraege],
        }

    @classmethod
    def from_json(cls, d: dict) -> Katalog:
        return cls(
            schule=str(d.get("schule", "") or ""),
            schuljahr=str(d.get("schuljahr", "") or ""),
            eintraege=[Eintrag.from_json(e) for e in d.get("eintraege", [])],
        )


# ── JSON-IO ──────────────────────────────────────────────────────────────────


def katalog_path() -> Path:
    """Kanonischer Pfad der Katalog-Datei: ``data/katalog.json``."""
    return paths.data_dir() / "katalog.json"


def load_katalog(path: Path | None = None, *, log: LogFn | None = None) -> Katalog:
    """Lädt den Katalog; leerer Katalog bei fehlender Datei.

    Bei kaputtem JSON (z. B. halbfertiger Schreibvorgang) wird ebenfalls ein
    leerer Katalog zurückgegeben + Warning geloggt — so bleibt die GUI bedienbar,
    statt beim Start abzustürzen. Strukturfehler (falsche Felder) hingegen werden
    weitergereicht (KeyError), da sie auf einen echten Datenfehler hindeuten.

    ``log`` (optional): Callback für die GUI-Logkonsole; ohne ihn fällt die
    Warnung auf ``logging`` zurück (s. :func:`_warn`).
    """
    if path is None:
        path = katalog_path()
    if not path.is_file():
        return Katalog(schule="", schuljahr="")
    try:
        return Katalog.from_json(json.loads(path.read_text(encoding="utf-8")))
    except json.JSONDecodeError as e:
        _warn(log, f"load_katalog: {path} ist kein gültiges JSON ({e.msg}) — leerer Katalog.")
        return Katalog(schule="", schuljahr="")


def save_katalog(katalog: Katalog, path: Path | None = None) -> Path:
    """Schreibt den Katalog als formatierte JSON-Datei; liefert den Pfad."""
    if path is None:
        path = katalog_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    _atomic_write_text(
        path,
        json.dumps(katalog.to_json(), indent=2, ensure_ascii=False) + "\n",
    )
    return path


# ── Import aus Excel + mappings ──────────────────────────────────────────────


def _norm_isbn(v: object) -> str:
    """ISBN auf Ziffern (+X) normieren — '978-3-12-…' → '978312…'."""
    return re.sub(r"[^0-9Xx]", "", str(v)) if v is not None else ""


def _zu_bestellen_lookup(wb: Workbook) -> dict[str, tuple[str, str, float]]:
    """Liest Sheet 'zu Bestellen': isbn_norm → (titel, verlag, neupreis).

    Spalten (1-basiert): B=Jahrgang, C=Fach, E=Verfasser/Titel, F=Verlag,
    G=ISBN, H=Einzelpreis. Nur Bücher mit Nachbestellbedarf liegen dort —
    fehlende ISBNs bleiben unangereichert (Titel/Verlag/Preis leer).
    """
    if "zu Bestellen" not in wb.sheetnames:
        return {}
    ws = wb["zu Bestellen"]
    out: dict[str, tuple[str, str, float]] = {}
    for row in range(2, ws.max_row + 1):
        isbn_norm = _norm_isbn(ws.cell(row, 7).value)  # G
        if not isbn_norm:
            continue
        titel = str(ws.cell(row, 5).value or "")  # E
        verlag = str(ws.cell(row, 6).value or "")  # F
        try:
            preis = float(ws.cell(row, 8).value or 0)  # H
        except (ValueError, TypeError):
            preis = 0.0
        out[isbn_norm] = (titel, verlag, preis)
    return out


def _grades_for_cell(ws, row: int, col: int) -> tuple[int, int] | None:
    """Jahrgang-Spanne (von, bis) einer Zelle aus ihrem (evtl. merged) Zeilenbereich.

    Liefert None, wenn der Bereich keine Jahrgang-Zeile enthält.
    """
    min_row, max_row = _merged_span(ws, row, col)
    grades = [
        extract_grade(ws, r)
        for r in range(min_row, max_row + 1)
        if classify_row(ws, r) == "jahrgang"
    ]
    grades = [g for g in grades if g is not None]
    if not grades:
        return None
    return min(grades), max(grades)


def _collect_row_types(ws) -> tuple[list[int], list[int]]:
    """Liefert (fach_rows, zustand_rows) aufsteigend."""
    fach_rows: list[int] = []
    zustand_rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        rt = classify_row(ws, row)
        if rt == "fach":
            fach_rows.append(row)
        elif rt == "zustand":
            zustand_rows.append(row)
    return fach_rows, zustand_rows


def import_from_excel(
    excel_path: Path | str,
    mappings: list[dict],
    *,
    sheet_name: str | None = None,
    schule: str = "",
    schuljahr: str = "",
    log: LogFn | None = None,
) -> Katalog:
    """Reverse-Import: Excel-Layout + config.json:mappings → Katalog.

    Für jeden mappings-Eintrag (bevorzugt ``bestand_cell``, sonst ``angemeldet_cell``)
    wird die Zelle in (Fach, Jahrgang von/bis, MJB) aufgelöst. Titel/Verlag/Neupreis
    werden — sofern vorhanden — aus dem Sheet 'zu Bestellen' per ISBN angereichert.
    Dedup nach (fach, hint, von, bis, isbn).

    ``log`` (optional): Callback für die GUI-Logkonsole für übersprungene
    Mapping-Einträge (s. :func:`_warn`).
    """
    wb = load_workbook(str(excel_path), data_only=False)
    sheet = sheet_name or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} fehlt in {excel_path}.")
    ws = wb[sheet]
    fach_rows, _zustand_rows = _collect_row_types(ws)
    enrich = _zu_bestellen_lookup(wb)

    seen: set[tuple[str, str | None, int, int, str]] = set()
    eintraege: list[Eintrag] = []
    for m in mappings:
        isbn = str(m.get("isbn", "") or "")
        if not isbn:
            continue
        cell = m.get("bestand_cell") or m.get("angemeldet_cell")
        if not cell:
            continue
        try:
            row, col = coordinate_to_tuple(str(cell))
        except (ValueError, IndexError, UnboundLocalError) as exc:
            # Eine kaputte Zellreferenz in einem einzelnen Mapping-Eintrag soll
            # nicht den ganzen Import abbrechen — nur dieses Mapping überspringen.
            _warn(
                log,
                f"import_from_excel: ungültige Zellreferenz {cell!r} ({exc}) — übersprungen.",
            )
            continue
        ar, ac = resolve_anchor(ws, row, col)
        span = _grades_for_cell(ws, ar, ac)
        if span is None:
            continue
        von, bis = span
        fach_val = find_fach_for_col(ws, fach_rows, ac)
        if fach_val is None:
            continue
        subject, hint = strip_hint(fach_val)
        key = (subject, hint, von, bis, isbn)
        if key in seen:
            continue
        seen.add(key)
        titel, verlag, preis = enrich.get(_norm_isbn(isbn), ("", "", 0.0))
        eintraege.append(
            Eintrag(
                fach=subject,
                jahrgang_von=von,
                jahrgang_bis=bis,
                isbn=isbn,
                hint=hint,
                titel=titel,
                verlag=verlag,
                neupreis=preis,
            )
        )

    # Stabile Reihenfolge: Fach, Jahrgang_von, Jahrgang_bis, Hint.
    eintraege.sort(key=lambda e: (e.fach, e.jahrgang_von, e.jahrgang_bis, e.hint or ""))
    return Katalog(schule=schule, schuljahr=schuljahr, eintraege=eintraege)


# ── match_overrides-Sync ─────────────────────────────────────────────────────


def catalog_to_overrides(katalog: Katalog, *, log: LogFn | None = None) -> dict[str, str]:
    """Leitet ``match_overrides`` aus dem Katalog ab (deterministisch, sortiert).

    Key-Format wie update_bestand_auto.py: ``f"{grade}|{fach}|{hint or ''}"``.
    Für Mehrjahresbände wird ein Key **pro** Jahrgang im Bereich [von..bis] emittiert
    — so greift der Override auf jeder Jahrgang-Zeile, die das Buch abdeckt.

    Bei überlappenden Einträgen (gleicher Key, unterschiedliche ISBN) gewinnt der
    letzte Eintrag — die Kollision wird aber geloggt, statt stillschweigend zu
    überschreiben (früher ein lautloser Datenverlust).

    ``log`` (optional): Callback für die GUI-Logkonsole (s. :func:`_warn`).
    """
    out: dict[str, str] = {}
    for e in katalog.eintraege:
        for grade in range(e.jahrgang_von, e.jahrgang_bis + 1):
            key = f"{grade}|{e.fach}|{e.hint or ''}"
            if key in out and out[key] != e.isbn:
                _warn(
                    log,
                    f"catalog_to_overrides: Key-Kollision für {key!r} — "
                    f"ISBN {e.isbn} überschreibt {out[key]} (letzter gewinnt).",
                )
            out[key] = e.isbn
    return dict(sorted(out.items()))


# ── Excel aus Vorlage (Mappings-only) ────────────────────────────────────────


def _mappings_for_layout(
    ws, katalog: Katalog
) -> tuple[list[dict], list[str]]:
    """Ordnet Katalog-Einträge den Layout-Slots der Vorlage zu.

    Liefert (mappings, unmatched). ``mappings``: Liste ``{isbn, bestand_cell,
    angemeldet_cell?}``. ``unmatched``: Beschreibungen der Einträge ohne Slot.
    """
    fach_rows, zustand_rows = _collect_row_types(ws)

    # grade → Zeile.
    grade_to_row: dict[int, int] = {}
    for row in range(1, ws.max_row + 1):
        if classify_row(ws, row) == "jahrgang":
            g = extract_grade(ws, row)
            if g is not None:
                grade_to_row[g] = row

    # (subject, hint) → {zustand_norm: col}.
    cols: dict[tuple[str, str | None], dict[str, int]] = {}
    for col in range(2, ws.max_column + 1):
        fach_val = find_fach_for_col(ws, fach_rows, col)
        if fach_val is None:
            continue
        zustand_val = find_zustand_for_col(ws, zustand_rows, col)
        if zustand_val is None:
            continue
        subject, hint = strip_hint(fach_val)
        cols.setdefault((subject, hint), {})[zustand_val.strip().lower()] = col

    mappings: list[dict] = []
    unmatched: list[str] = []

    def _unmatched(e: Eintrag, why: str) -> str:
        return f"{e.fach} ({e.hint or '-'}) Jg.{e.jahrgang_von}-{e.jahrgang_bis} [{e.isbn}] — {why}"

    for e in katalog.eintraege:
        slot = cols.get((e.fach, e.hint))
        if slot is None:
            unmatched.append(_unmatched(e, "kein Layout-Slot"))
            continue
        rows = [grade_to_row.get(g) for g in range(e.jahrgang_von, e.jahrgang_bis + 1)]
        if not rows:
            # jahrgang_von > jahrgang_bis ⇒ range() ist leer ⇒ any([]) wäre False,
            # der min()-Aufruf unten würde also auf einer leeren Sequenz crashen.
            # Eintrag.__post_init__ tauscht von/bis normalerweise schon — dieser
            # Guard bleibt trotzdem als defensive Absicherung bestehen.
            unmatched.append(_unmatched(e, "Jahrgang-Bereich leer/invertiert"))
            continue
        if any(r is None for r in rows):
            unmatched.append(_unmatched(e, "Jahrgang nicht im Layout"))
            continue
        min_row = min(rows)  # type: ignore[arg-type]

        entry: dict = {"isbn": e.isbn}
        bestand_col = slot.get("bestand")
        if bestand_col is not None:
            ar, ac = resolve_anchor(ws, min_row, bestand_col)
            entry["bestand_cell"] = f"{get_column_letter(ac)}{ar}"
        angemeldet_col = slot.get("angemeldet")
        if angemeldet_col is not None:
            ar, ac = resolve_anchor(ws, min_row, angemeldet_col)
            entry["angemeldet_cell"] = f"{get_column_letter(ac)}{ar}"
        if "bestand_cell" not in entry and "angemeldet_cell" not in entry:
            unmatched.append(_unmatched(e, "keine Bestand/Angemeldet-Spalte"))
            continue
        mappings.append(entry)

    return mappings, unmatched


def render_excel(
    template: Path | str,
    katalog: Katalog,
    out_path: Path | str,
    *,
    sheet_name: str | None = None,
    safety_stock: int = 5,
) -> tuple[Path, list[str]]:
    """Kopiert die Vorlage nach ``out_path`` und schreibt eine config.json daneben.

    Mappings-only (Nutzerentscheid 2026-08-12): das Layout wird **nicht** neu
    aufgebaut — die Vorlage bleibt unverändert. Aus dem Vorlagen-Layout + Katalog
    werden ``mappings`` (ISBN→Zelle) abgeleitet und neben der Ausgabe als
    ``config.json`` gespeichert (plus ``match_overrides`` aus dem Katalog).

    Rückgabe: (config_json_path, unmatched_beschreibungen).
    """
    out_path = Path(out_path)
    template = Path(template)
    shutil.copyfile(template, out_path)

    wb = load_workbook(str(out_path), data_only=False)
    sheet = sheet_name or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} fehlt in Vorlage {template}.")
    mappings, unmatched = _mappings_for_layout(wb[sheet], katalog)

    config = {
        "excel_file": out_path.name,
        "sheet_name": sheet,
        "safety_stock": safety_stock,
        "match_overrides": catalog_to_overrides(katalog),
        "mappings": mappings,
    }
    config_path = out_path.parent / "config.json"
    _atomic_write_text(
        config_path,
        json.dumps(config, indent=2, ensure_ascii=False) + "\n",
    )
    return config_path, unmatched
