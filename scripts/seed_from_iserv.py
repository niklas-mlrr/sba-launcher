#!/usr/bin/env python3
"""Einmaliger Katalog-Seed per IServ-GET (Phase 4, 2026-08-12).

Lokale Quellen (config.json:mappings) sind teils veraltet (19/71 zeigen auf keine
Jahrgang-Zeile). Dieser Seed nutzt stattdessen das **Matching aus update_bestand_auto**
(Bücherlisten + Serien per GET, ``match_book``) als autoritative Quelle und schreibt
einen vollständigen Katalog mit Titel/Verlag/Neupreis nach ``data/katalog.json``.
Zusätzlich wird die 2026er Excel als Vorlage nach ``templates/Bestand-Vorlage.xlsx``
kopiert.

Läuft im **ausleihe-api-.venv** (hat ``ausleihe`` + openpyxl + isbnlib + dotenv), NICHT
im Launcher-Venv (das ``ausleihe`` nicht enthält):

    ~/projects/sba/ausleihe-api/.venv/bin/python \\
        ~/projects/sba/sba-launcher/scripts/seed_from_iserv.py

Produktionsschutz: rein lesend (GET). Keine Buchungen, keine API-Writes, kein
``ALLOW_BOOKING``. Credentials aus ``ausleihe-api/.env`` (via dotenv).
"""

from __future__ import annotations

import sys
from pathlib import Path

# sba-launcher-Root (für ``core.catalog`` + ``core.paths``) + Bestand-Dir (für
# ``update_bestand_auto``) in den Pfad nehmen. Reihenfolge: Bestand zuerst,
# damit update_bestand_auto sein eigenes sys.path-Setup (_ROOT = parent.parent)
# ungestört vorfindet.
LAUNCHER = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(LAUNCHER))

from core import paths  # noqa: E402

BESTAND = paths.sibling("ausleihe-api") / "bestand- und nachbestellungen/New - API approach"
sys.path.insert(0, str(BESTAND))

import update_bestand_auto as auto  # noqa: E402
from ausleihe import AusleiheClient  # noqa: E402
from ausleihe.inventory_excel import match_book  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from core import catalog  # noqa: E402

EXCEL = BESTAND / "Bestand- und Nachbestellungsliste 2026.xlsx"
CONFIG = BESTAND / "config.json"
SCHULE = "TRG Osterode"
SCHULJAHR = "2026/2027"


def _to_float(value) -> float:
    """Wandelt einen Preis-Wert in float; Komma-Dezimaltrenner → Punkt.

    Fehlt oder unparsebar → 0.0 (sensible Default für fehlende Neupreise).
    """
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def main() -> None:
    import json

    config = json.loads(CONFIG.read_text(encoding="utf-8"))
    sheet = config.get("sheet_name") or "Bestand- und Nachbestellung"

    print("Verbinde mit IServ …")
    client = AusleiheClient()
    # ``get_current`` 404t auf diesem IServ (kein 'current'-Marker gesetzt) —
    # Schuljahr explizit verwenden (dasselbe, das der Bestand-Excel zugrunde liegt).
    sy_id = SCHULJAHR
    print(f"Schuljahr: {sy_id}")

    booklists = client.schoolyears.get_booklists(sy_id)
    booklists_by_grade = {bl["grade"]: bl for bl in booklists if bl.get("grade") is not None}
    print("Lade Serien-Daten …")
    series_data = auto.fetch_series_data(client)  # isbn -> {total, publisher, price, title}

    wb = load_workbook(str(EXCEL))
    ws = wb[sheet]
    fach_rows: list[int] = []
    zustand_rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        rt = auto.classify_row(ws, row)
        if rt == "fach":
            fach_rows.append(row)
        elif rt == "zustand":
            zustand_rows.append(row)

    grade_books_cache: dict[int, list[dict]] = {}
    seen: set[tuple[str, str | None, int, int, str]] = set()
    eintraege: list[catalog.Eintrag] = []

    for row in range(1, ws.max_row + 1):
        if auto.classify_row(ws, row) != "jahrgang":
            continue
        grade = auto.extract_grade(ws, row)
        if grade is None or not fach_rows:
            continue
        if grade not in grade_books_cache:
            bl = booklists_by_grade.get(grade)
            if bl:
                print(f"  Lade Bücherliste Jahrgang {grade} …")
                grade_books_cache[grade] = auto.load_grade_books(client, sy_id, bl["id"])
            else:
                grade_books_cache[grade] = []
        books = grade_books_cache[grade]

        for col in range(2, ws.max_column + 1):
            zustand_label = auto.find_zustand_for_col(ws, zustand_rows, col)
            if zustand_label is None:
                continue
            zustand_norm = zustand_label.strip().lower()
            # Eine Spalte pro (Fach, Buch) reicht — Bestand oder Angemeldet triggern.
            if zustand_norm not in ("bestand", "angemeldet"):
                continue
            fach_val = auto.find_fach_for_col(ws, fach_rows, col)
            if fach_val is None:
                continue
            subject, hint = auto.strip_hint(fach_val)
            match = match_book(
                books, subject, hint, hint_expansions=auto._HINT_EXPANSIONS
            )
            if match.book is None:
                continue
            isbn = match.book["isbn"]
            ar, ac = auto.resolve_anchor(ws, row, col)
            span = catalog._grades_for_cell(ws, ar, ac)
            if span is None:
                continue
            von, bis = span
            key = (subject, hint, von, bis, isbn)
            if key in seen:
                continue
            seen.add(key)
            sd = series_data.get(isbn, {})
            eintraege.append(
                catalog.Eintrag(
                    fach=subject,
                    jahrgang_von=von,
                    jahrgang_bis=bis,
                    isbn=isbn,
                    hint=hint,
                    titel=str(sd.get("title", "") or ""),
                    verlag=str(sd.get("publisher", "") or ""),
                    neupreis=_to_float(sd.get("price", 0)),
                )
            )

    eintraege.sort(key=lambda e: (e.fach, e.jahrgang_von, e.jahrgang_bis, e.hint or ""))
    katalog = catalog.Katalog(schule=SCHULE, schuljahr=SCHULJAHR, eintraege=eintraege)

    out_katalog = catalog.katalog_path()
    out_template = catalog.paths.templates_dir() / "Bestand-Vorlage.xlsx"
    catalog.save_katalog(katalog, out_katalog)
    out_template.parent.mkdir(parents=True, exist_ok=True)
    import shutil

    shutil.copyfile(EXCEL, out_template)

    mjb = [e for e in eintraege if e.mehrjahresband]
    ang = [e for e in eintraege if e.titel]
    print(f"\nFertig: {len(eintraege)} Einträge ({len(mjb)} MJB, {len(ang)} mit Titel).")
    print(f"  Katalog:   {out_katalog}")
    print(f"  Vorlage:   {out_template}")


if __name__ == "__main__":
    main()
